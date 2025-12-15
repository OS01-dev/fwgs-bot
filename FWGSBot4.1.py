"""
Unified Fine Wine & Good Spirits Bot (Complex Version C)
- Single-file implementation
- Watchlists, Global list, Hotlist, daily Excel report + comparison
- Hotlist triggered by Category -> 'whiskey-release', monitored every 60s
- Schedules: global refresh (5m), hot refresh (60s), daily report (9:05)
"""

"""
FWGS Bot - PostgreSQL Version
Complete foundation with all necessary imports and setup
"""

import os
import asyncio
import traceback
import requests
import re
import aiohttp
from datetime import datetime, timedelta, time as datetime_time
from dotenv import load_dotenv
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from telegram import LabeledPrice, PreCheckoutQuery
from datetime import datetime, timedelta
from telegram import LabeledPrice, PreCheckoutQuery
from datetime import time
import warnings

# Telegram imports
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
    PreCheckoutQueryHandler
)

from telegram.warnings import PTBUserWarning  # ← Add this



# Scheduler imports
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# PostgreSQL imports
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
import aiohttp

# -------------------------
# Configuration - Set your credentials here
# -------------------------
BOT_TOKEN = os.getenv("BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
OWNER_CHAT_ID = os.getenv("OWNER_CHAT_ID")

if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN environment variable not set!")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL environment variable not set!")

# Validate configuration
if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
    raise ValueError("Please set BOT_TOKEN in the script!")
if not DATABASE_URL or DATABASE_URL == "YOUR_DATABASE_URL_HERE":
    raise ValueError("Please set DATABASE_URL in the script!")

# Convert OWNER_CHAT_ID to int if provided
try:
    OWNER_CHAT_ID = int(OWNER_CHAT_ID) if str(OWNER_CHAT_ID).strip() != "" else None
except Exception:
    OWNER_CHAT_ID = None

# -------------------------
# Configuration Constants
# -------------------------
BASE_URL = "https://www.finewineandgoodspirits.com"
REPORT_PREFIX = "product_report_"

# API endpoints
FWGS_STOCKSTATUS_URL = (
    "https://www.finewineandgoodspirits.com/ccstore/v1/stockStatus"
    "?actualStockStatus=true&expandStockDetails=true&products=repositoryId:{pid}&locationIds={store}"
)
FWGS_LOCATION_URL = "https://www.finewineandgoodspirits.com/ccstore/v1/locations/{store}"
PRODUCT_BASE_URL = "https://www.finewineandgoodspirits.com/ccstore/v1/products"

# Business hours for notifications (optional - can adjust later)
from datetime import time
BUSINESS_START = time(8, 0)      # 08:00 local
BUSINESS_END = time(21, 0)       # 21:00 local
POLL_INTERVAL_SECONDS = 30 * 60  # 30 minutes
TRIAL_DAYS = 14  # 2 weeks free trial
SUBSCRIPTION_PRICE_STARS = 300  # ~$5 (1 Star ≈ $0.015-0.02)
SUBSCRIPTION_DURATION_DAYS = 30  # 30 days per payment

# -------------------------
# Logging Helper
# -------------------------
def log(message):
    """Simple logging with timestamp."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Log errors caused by Updates and Jobs."""
    log(f"❌ Exception while handling an update: {context.error}")
    log(f"Traceback: {traceback.format_exc()}")

# ============================================================
# DATABASE CONNECTION POOL & SETUP
# ============================================================
connection_pool = None

def init_connection_pool():
    """Initialize the connection pool on startup."""
    global connection_pool
    try:
        connection_pool = psycopg2.pool.SimpleConnectionPool(
            minconn=2,
            maxconn=20,  # Increased for better concurrency
            dsn=DATABASE_URL,
            # Add connection timeout and keepalive
            connect_timeout=10,
            keepalives=1,
            keepalives_idle=30,
            keepalives_interval=10,
            keepalives_count=5
        )
        log("✅ Connection pool created successfully")
    except Exception as e:
        log(f"❌ Error creating connection pool: {e}")
        raise

def get_db():
    """Get a connection from the pool."""
    if connection_pool is None:
        init_connection_pool()
    return connection_pool.getconn()

def return_db(conn):
    """Return a connection to the pool."""
    if connection_pool:
        connection_pool.putconn(conn)

def close_pool():
    """Close all connections in the pool (call on shutdown)."""
    if connection_pool:
        connection_pool.closeall()
        log("✅ Connection pool closed")

def init_db():
    """Create all tables expected by the application."""
    conn = get_db()
    try:
        cur = conn.cursor()
        
        # USERS
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id TEXT PRIMARY KEY,
                full_name TEXT,
                is_admin BOOLEAN DEFAULT FALSE,
                is_subscribed BOOLEAN DEFAULT FALSE,
                subscription_expiry TIMESTAMP,
                joined TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # WATCHLIST
        cur.execute("""
            CREATE TABLE IF NOT EXISTS watchlist (
                user_id TEXT NOT NULL,
                user_name TEXT,
                product_id TEXT NOT NULL,
                product_name TEXT,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, product_id)
            );
        """)
        
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_watchlist_product 
            ON watchlist(product_id);
        """)
        
        # STORES
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stores (
                user_id TEXT NOT NULL,
                store_id TEXT NOT NULL,
                city TEXT,
                address1 TEXT,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, store_id)
            );
        """)
        
        # GLOBAL_PRODUCTS
        cur.execute("""
            CREATE TABLE IF NOT EXISTS global_products (
                product_id TEXT PRIMARY KEY,
                name TEXT,
                category TEXT,
                in_stock BOOLEAN DEFAULT FALSE,
                price TEXT,
                active BOOLEAN DEFAULT FALSE,
                allocated TEXT,
                lottery TEXT,
                order_limit TEXT,
                product_full_url TEXT,
                thumbnail_url TEXT,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_global_products_active 
            ON global_products(active) WHERE active = TRUE;
        """)
        
        # PRODUCT_CACHE
        cur.execute("""
            CREATE TABLE IF NOT EXISTS product_cache (
                product_id TEXT PRIMARY KEY,
                name TEXT,
                category TEXT,
                last_qty INTEGER,
                updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # PRODUCT_QUANTITY_CACHE
        cur.execute("""
            CREATE TABLE IF NOT EXISTS product_quantity_cache (
                product_id TEXT NOT NULL,
                store_id TEXT NOT NULL,
                last_qty INTEGER DEFAULT 0,
                last_checked TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (product_id, store_id)
            );
        """)
        
        # SUBSCRIPTIONS
        cur.execute("""
            CREATE TABLE IF NOT EXISTS subscriptions (
                user_id TEXT PRIMARY KEY,
                stripe_customer_id TEXT,
                stripe_subscription_id TEXT,
                start_date TIMESTAMP,
                trial_end TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        
        # STORE_STOCK
        cur.execute("""
            CREATE TABLE IF NOT EXISTS store_stock (
                product_id TEXT NOT NULL,
                store_id TEXT NOT NULL,
                quantity INTEGER DEFAULT 0,
                last_checked TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (product_id, store_id)
            );
        """)
        # FWGS_STORES - All stores in Pennsylvania
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fwgs_stores (
                store_id TEXT PRIMARY KEY,
                city TEXT,
                address1 TEXT,
                state TEXT DEFAULT 'PA',
                zip_code TEXT,
                phone TEXT,
                latitude NUMERIC,
                longitude NUMERIC,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_fwgs_stores_city 
            ON fwgs_stores(city);
        """)
        
        conn.commit()
        log("✅ Database initialized successfully")
        
    except Exception as e:
        conn.rollback()
        log(f"❌ Error initializing database: {e}")
        raise
    finally:
        cur.close()
        return_db(conn)

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def normalize_pid(raw_pid):
    """Normalize product ID (implement your validation logic)."""
    # Add your normalization logic here
    # For now, just strip whitespace and return
    cleaned = str(raw_pid).strip()
    return cleaned if cleaned else None

def to_number(v):
    """Convert various formats to number for Excel processing."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return 0
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    s = re.sub(r"[^\d\.\-eE]", "", s)
    try:
        if "." in s or "e" in s or "E" in s:
            return float(s)
        return int(float(s))
    except Exception:
        return 0

def normalize_for_comparison(val, col_name):
    """Normalize values for comparison between reports."""
    if val in [None, "N/A", "", float('nan')] or (isinstance(val, float) and pd.isna(val)):
        return None
    if col_name in ["InStock", "OdrLmt", "Price"]:
        try:
            return float(val)
        except Exception:
            return None
    if col_name in ["Active", "Allocated", "Lottery"]:
        if isinstance(val, bool):
            return "Y" if val else "N"
        val_str = str(val).strip().upper()
        if val_str in ["Y", "YES", "TRUE"]:
            return "Y"
        elif val_str in ["N", "NO", "FALSE"]:
            return "N"
        return val_str
    return str(val).strip()

# ============================================================
# API FUNCTIONS
# ============================================================

BASE_URL = "https://www.finewineandgoodspirits.com"

def get_product_info(pid):
    """Fetch product details from the FWGS API."""
    headers = {"User-Agent": "Mozilla/5.0"}
    product_url = f"{BASE_URL}/ccstore/v1/products/{pid}"
    stock_url = (
        f"{BASE_URL}/ccstore/v1/stockStatus"
        f"?actualStockStatus=true&expandStockDetails=true&products=repositoryId%3A{pid}&locationIds=9650"
    )
    
    try:
        res = requests.get(product_url, headers=headers, timeout=10)
        if res.status_code >= 400:
            log(f"❌ Invalid product ID {pid} (HTTP {res.status_code})")
            return None
        
        data = res.json()
        
        parent_categories = data.get("parentCategories", [])
        parent_category = ", ".join(
            cat.get("displayName", cat.get("repositoryId", "N/A"))
            for cat in parent_categories
        ) if parent_categories else "N/A"
        
        active = data.get("active", "N/A")
        display_name = data.get("displayName", "Unknown")
        highly_allocated = data.get("b2c_highlyAllocatedProduct", "N/A")
        lottery_product = data.get("b2c_lotteryProduct", "N/A")
        repo_id = data.get("repositoryId", pid)
        order_limit = data.get("b2c_limitPerOrder", "N/A")
        route = data.get("route", "")
        product_full_url = f"{BASE_URL}{route}" if route else f"{BASE_URL}/product/{pid}"
        
        thumb_path = data.get("primarySmallImageURL")
        thumbnail_url = f"{BASE_URL}{thumb_path}" if thumb_path else None
        
        list_price = data.get("listPrice", "N/A")
        if isinstance(list_price, dict):
            list_price = list_price.get("value", "N/A")
        
        # Fetch stock
        try:
            stock_res = requests.get(stock_url, headers=headers, timeout=10)
            if stock_res.status_code < 400:
                stock_data = stock_res.json()
                stock_info = stock_data.get("items", [])[0] if stock_data.get("items") else {}
                quantity = stock_info.get("inStockQuantity", 0)
                if not str(quantity).isdigit():
                    quantity = 0
            else:
                quantity = 0
        except Exception:
            quantity = 0
        
        return {
            "ProductID": str(repo_id),
            "Name": display_name,
            "Category": parent_category,
            "Active": str(active),
            "InStock": int(quantity),
            "Allocated": highly_allocated,
            "Lottery": lottery_product,
            "Price": str(list_price),
            "OdrLmt": order_limit,
            "Thumbnail": thumbnail_url,
            "product_full_url": product_full_url
        }
    except Exception as e:
        log(f"❌ Error fetching product {pid}: {e}")
        return None

# ============================================================
# DATABASE OPERATIONS - WATCHLIST
# ============================================================

def add_to_watchlist(user_id, user_name, product_id, product_name):
    """Add a product to a user's watchlist."""
    query = """
        INSERT INTO watchlist (user_id, user_name, product_id, product_name)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (user_id, product_id) DO NOTHING
        RETURNING product_id;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, user_name, product_id, product_name))
        result = cur.fetchone()
        conn.commit()
        cur.close()
        return result is not None
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error adding to watchlist: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

def remove_from_watchlist(user_id, product_id):
    """Remove a product from a user's watchlist."""
    query = """
        DELETE FROM watchlist 
        WHERE user_id = %s AND product_id = %s
        RETURNING product_name;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, product_id))
        result = cur.fetchone()
        conn.commit()
        cur.close()
        return result[0] if result else None  # Returns product name if deleted, None if not found
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error removing from watchlist: {e}")
        return None
    finally:
        if conn:
            return_db(conn)

def is_in_watchlist(user_id, product_id):
    """Check if product is in user's watchlist."""
    query = "SELECT 1 FROM watchlist WHERE user_id = %s AND product_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, product_id))
        result = cur.fetchone()
        cur.close()
        return result is not None
    except Exception as e:
        log(f"❌ Error checking watchlist: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

def get_user_watchlist(user_id):
    """Get all products in a user's watchlist."""
    query = """
        SELECT product_id, product_name 
        FROM watchlist 
        WHERE user_id = %s
        ORDER BY added_at DESC;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id,))
        results = cur.fetchall()
        cur.close()
        return results  # Returns list of tuples: [(product_id, product_name), ...]
    except Exception as e:
        log(f"❌ Error getting watchlist: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

# ============================================================
# DATABASE OPERATIONS - STORES
# ============================================================

def add_user_store(user_id, store_id, city, address):
    """Add a store to a user's tracked stores."""
    query = """
        INSERT INTO stores (user_id, store_id, city, address1)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (user_id, store_id) DO NOTHING
        RETURNING store_id;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, store_id, city, address))
        result = cur.fetchone()
        conn.commit()
        cur.close()
        return result is not None  # True if added, False if already existed
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error adding store: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

def remove_user_store(user_id, store_id):
    """Remove a store from a user's tracked stores."""
    query = """
        DELETE FROM stores 
        WHERE user_id = %s AND store_id = %s
        RETURNING city;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, store_id))
        result = cur.fetchone()
        conn.commit()
        cur.close()
        return result[0] if result else None  # Returns city if deleted, None if not found
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error removing store: {e}")
        return None
    finally:
        if conn:
            return_db(conn)

def get_user_stores(user_id):
    """Get all stores tracked by a user."""
    query = """
        SELECT store_id, city, address1 
        FROM stores 
        WHERE user_id = %s
        ORDER BY added_at ASC;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id,))
        results = cur.fetchall()
        cur.close()
        return results  # Returns list of tuples: [(store_id, city, address), ...]
    except Exception as e:
        log(f"❌ Error getting user stores: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

def is_store_tracked(user_id, store_id):
    """Check if a store is already tracked by user."""
    query = "SELECT 1 FROM stores WHERE user_id = %s AND store_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, store_id))
        result = cur.fetchone()
        cur.close()
        return result is not None
    except Exception as e:
        log(f"❌ Error checking store: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

def add_to_global_products(product_info):
    """Add or update a product in global products table."""
    query = """
        INSERT INTO global_products (
            product_id, name, category, in_stock, price, active,
            allocated, lottery, order_limit, product_full_url, thumbnail_url, last_updated
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id) DO UPDATE SET
            name = EXCLUDED.name,
            category = EXCLUDED.category,
            in_stock = EXCLUDED.in_stock,
            price = EXCLUDED.price,
            active = EXCLUDED.active,
            allocated = EXCLUDED.allocated,
            lottery = EXCLUDED.lottery,
            order_limit = EXCLUDED.order_limit,
            product_full_url = EXCLUDED.product_full_url,
            thumbnail_url = EXCLUDED.thumbnail_url,
            last_updated = CURRENT_TIMESTAMP;
    """
    
    active_bool = product_info.get("Active", "false").lower() in ("true", "yes", "1")
    
    params = (
        product_info["ProductID"],
        product_info["Name"],
        product_info["Category"],
        product_info["InStock"] > 0,
        product_info["Price"],
        active_bool,
        product_info["Allocated"],
        product_info["Lottery"],
        product_info["OdrLmt"],
        product_info["product_full_url"],
        product_info["Thumbnail"]
    )
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, params)
        conn.commit()
        cur.close()
        return True
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error adding to global products: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

def is_in_global_products(product_id):
    """Check if product exists in global products."""
    query = "SELECT 1 FROM global_products WHERE product_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        result = cur.fetchone()
        cur.close()
        return result is not None
    except Exception as e:
        log(f"❌ Error checking global products: {e}")
        return False
    finally:
        if conn:
            return_db(conn)

# ============================================================
# TELEGRAM HANDLERS 
# ============================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Modified /start to create user with trial."""
    user = update.effective_user
    user_id = str(user.id)
    full_name = user.full_name or user.username or "Unknown"
    
    # Create user with trial if new
    create_or_update_user(user_id, full_name)
    
    # Get subscription status
    status = get_user_subscription_status(user_id)
    
    welcome_msg = (
        "🥃 <b>Welcome to the Fine Wine & Good Spirits Bot!</b>\n\n"
        "📌 This bot tracks product availability and sends real-time alerts.\n\n"
    )
    
    if status and status["is_admin"]:
        welcome_msg += "👑 <b>Admin Access</b> - Full access enabled\n\n"
    elif status and status["trial_active"]:
        days_left = (status["expiry"] - datetime.now()).days
        welcome_msg += f"🎁 <b>Free Trial Active</b> - {days_left} days remaining\n\n"
    elif status and status["is_active"]:
        days_left = (status["expiry"] - datetime.now()).days
        welcome_msg += f"✅ <b>Subscribed</b> - {days_left} days remaining\n\n"
    else:
        welcome_msg += "⭐ <b>Trial Expired</b> - Use /subscribe to continue\n\n"
    
    welcome_msg += (
        "📌 There are two lists: a Global list and your Custom list.\n"
        "• The Global list is sent each morning at 9 AM.\n"
        "• Your Custom list is used for alerts — both online and in-store.\n"
        "• Make sure to add products AND your nearby stores.\n\n"
        "Commands:\n"
        "/add <i>ProductID</i> - Add product to your custom watchlist.\n"
        "/remove <i>ProductID</i> - Remove product from your custom watchlist.\n"
        "/watchlist - Show your watchlist.\n"
        "/global - Show global product list.\n"
        "/lastreport - Send the last report.\n"
        "/addstore <i>StoreID</i> - Adds a store to your list.\n"
        "/removestore <i>StoreID</i> - Remove a store from your list.\n"
        "/mystores - Show my store list.\n"
        "/statestock <i>ProductID</i> - Shows all stock statewide.\n"
        "/subscribe - Subscribe for $5/month.\n"
        "/messageadmin <i>Message</i> - Send a message to admin.\n"
        "💡 You can also send a product ID directly to get its latest info and stock, both e-Commerce and MyStore list.\n"
    )
    
    await update.message.reply_text(welcome_msg, parse_mode="HTML")

async def add_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Add products to user's watchlist."""
    user = update.effective_user
    user_id = str(user.id)
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        if reason == "no_account":
            await update.message.reply_text(
                "👋 Welcome! Send /start to begin your free trial."
            )
        else:
            await update.message.reply_text(
                "⭐ <b>Premium Feature</b>\n\n"
                "Subscribe to add products to your watchlist:\n/subscribe",
                parse_mode="HTML"
            )
        return
    
    username = user.full_name or user.username or "Unknown"
    
    if not context.args:
        await update.message.reply_text("Usage: /add <ProductID> [ProductID ...]")
        return
    
    raw_ids = [pid.strip() for pid in context.args if pid.strip()]
    if not raw_ids:
        await update.message.reply_text("⚠️ No valid product IDs provided.")
        return
    
    added = []
    already_in = []
    invalid = []
    
    for raw_pid in raw_ids:
        pid = normalize_pid(raw_pid)
        if pid is None:
            invalid.append(raw_pid)
            continue
        
        info = get_product_info(pid)
        if not info:
            invalid.append(raw_pid)
            continue
        
        if is_in_watchlist(user_id, pid):
            already_in.append(pid)
        else:
            if add_to_watchlist(user_id, username, pid, info.get("Name", "Unknown Product")):
                added.append((pid, info.get("Name", "Unknown Product")))
        
        if not is_in_global_products(pid):
            add_to_global_products(info)
    
    msg_lines = []
    if added:
        msg_lines.append(
            f"✅ Added: {', '.join([f'{name} (ID {pid})' for pid, name in added])}"
        )
    if already_in:
        msg_lines.append(f"ℹ Already in watchlist: {', '.join(already_in)}")
    if invalid:
        msg_lines.append(f"❌ Invalid product IDs: {', '.join(invalid)}")
    
    await update.message.reply_text("\n".join(msg_lines))

async def remove_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove a product from user's watchlist - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to manage your watchlist: /subscribe"
        )
        return
    
    if not context.args:
        await update.message.reply_text("Usage: /remove <ProductID>")
        return
    
    pid = context.args[0].strip()
    removed_name = remove_from_watchlist(user_id, pid)
    
    if removed_name:
        await update.message.reply_text(f"🗑 Removed {pid} – {removed_name} from your watchlist.")
    else:
        await update.message.reply_text(f"⚠️ {pid} not found on your watchlist.")

async def show_watchlist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show user's watchlist - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to view your watchlist: /subscribe"
        )
        return
    
    watchlist = get_user_watchlist(user_id)
    
    if not watchlist:
        await update.message.reply_text("👁 Your watchlist is empty.")
        return
    
    sorted_watchlist = sorted(watchlist, key=lambda p: p[1].lower())
    lines = [f"{pid} – {name}" for pid, name in sorted_watchlist]
    
    MAX_LINES = 25
    for i in range(0, len(lines), MAX_LINES):
        chunk = lines[i:i+MAX_LINES]
        await update.message.reply_text("👁 Your watchlist:\n" + "\n".join(chunk))

async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show bot status."""
    await update.message.reply_text("✅ Bot online. Scheduler running.")

async def kill_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Emergency shutdown - owner only."""
    # Permission check
    if str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ You are not authorized to shut down the bot.")
        return
    
    log("🛑 Bot killed by owner")
    
    msg = "🛑 Shutting down bot...\n"
    
    # Stop scheduler if running
    # Note: scheduler is a local variable in runner(), but we can stop the app
    msg += "✅ Stopping all services\n"
    
    await update.message.reply_text(msg)
    
    # Clean shutdown
    try:
        await context.application.updater.stop()
        await context.application.stop()
        await context.application.shutdown()
        close_pool()  # Close database connections
        log("✅ Clean shutdown complete")
    except Exception as e:
        log(f"⚠️ Error during shutdown: {e}")
    
    # Force exit
    import sys
    sys.exit(0)

async def show_global_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show the global product list."""
    # Get all products from global_products table
    query = """
        SELECT product_id, name 
        FROM global_products 
        ORDER BY LOWER(name);
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        products = cur.fetchall()
        cur.close()
    except Exception as e:
        log(f"❌ Error getting global products: {e}")
        await update.message.reply_text("⚠️ Error retrieving global product list.")
        return
    finally:
        if conn:
            return_db(conn)
    
    if not products:
        await update.message.reply_text("🌐 Global product list is empty.")
        return
    
    # Format as "ProductID – Name"
    msg_lines = [f"{pid} – {name}" for pid, name in products]
    
    # Send in chunks of 50 to avoid message length limits
    CHUNK_SIZE = 50
    for i in range(0, len(msg_lines), CHUNK_SIZE):
        chunk = msg_lines[i:i+CHUNK_SIZE]
        await update.message.reply_text("🌐 Global products:\n" + "\n".join(chunk))

async def send_global_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Report handler - Coming soon!")

async def addstore_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Add a store to user's tracked stores - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to track stores: /subscribe"
        )
        return
    
    if not context.args:
        await update.message.reply_text("Usage: /addstore <store_number>")
        return
    
    try:
        store = int(context.args[0])
        store_id = str(store)
    except ValueError:
        await update.message.reply_text("Store number must be numeric.")
        return
    
    if is_store_tracked(user_id, store_id):
        await update.message.reply_text(f"Store {store} is already in your tracked stores.")
        return
    
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
    
    import aiohttp
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(
                FWGS_LOCATION_URL.format(store=store),
                timeout=aiohttp.ClientTimeout(total=10),
                headers=headers
            ) as resp:
                if resp.status != 200:
                    await update.message.reply_text(
                        f"Failed to fetch store info for {store}. Status {resp.status}"
                    )
                    return
                data = await resp.json()
        except Exception as e:
            await update.message.reply_text(f"Error fetching store info for {store}: {e}")
            return
    
    city = data.get("city", "Unknown")
    address = data.get("address1", "Unknown")
    
    if add_user_store(user_id, store_id, city, address):
        await update.message.reply_text(
            f"✅ Added store {store}: {city}, {address} to your tracked stores."
        )
    else:
        await update.message.reply_text(f"⚠️ Failed to add store {store}.")

async def removestore_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove a store from user's tracked stores - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to manage your stores: /subscribe"
        )
        return
    
    if not context.args:
        await update.message.reply_text("Usage: /removestore <store_number>")
        return
    
    try:
        store_id = str(int(context.args[0]))
    except ValueError:
        await update.message.reply_text("Store number must be numeric.")
        return
    
    removed_city = remove_user_store(user_id, store_id)
    
    if removed_city:
        await update.message.reply_text(
            f"🗑 Removed store {store_id} ({removed_city}) from your tracked stores."
        )
    else:
        await update.message.reply_text(f"⚠️ Store {store_id} is not in your tracked stores.")


async def mystores_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """List all stores the user is tracking - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to view your stores: /subscribe"
        )
        return
    
    user_stores = get_user_stores(user_id)
    
    if not user_stores:
        await update.message.reply_text(
            "You have no tracked stores. Add one with /addstore <store_number>."
        )
        return
    
    lines = []
    for store_id, city, address in user_stores:
        addr_line = ", ".join(filter(None, [address, city]))
        lines.append(f"{store_id} — {addr_line}")
    
    await update.message.reply_text("🏪 Your tracked stores:\n" + "\n".join(lines))


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle direct product ID messages - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to check product info: /subscribe"
        )
        return
    
    import aiohttp
    
    text = update.message.text.strip()
    
    if not text.isdigit():
        await update.message.reply_text("Please send a numeric product ID or use a command.")
        return
    
    pid = text
    await update.message.reply_text(f"🔎 Fetching info for {pid}...")
    
    info = get_product_info(pid)
    if not info:
        await update.message.reply_text("⚠️ Could not retrieve product info.")
        return
    
    msg = (
        f"{info.get('Name','')}\n"
        f"Category: {info.get('Category','')}\n"
        f"E-Commerce Stock: {info.get('InStock','')}\n"
        f"Active: {info.get('Active','')}\n"
        f"Price: ${info.get('Price','')}\n"
        f"Product ID: {info.get('ProductID','')}\n"
        f"Lottery: {info.get('Lottery','')}\n"
        f"Order Limit: {info.get('OdrLmt','')}"
    )
    
    user_stores = get_user_stores(user_id)
    
    if user_stores:
        headers = {"User-Agent": "Mozilla/5.0"}
        location_ids_str = ",".join([store_id for store_id, _, _ in user_stores])
        url = FWGS_STOCKSTATUS_URL.format(pid=pid, store=location_ids_str)
        
        async with aiohttp.ClientSession() as session:
            try:
                async with session.get(
                    url, 
                    timeout=aiohttp.ClientTimeout(total=10), 
                    headers=headers
                ) as resp:
                    if resp.status == 200:
                        data = await resp.json(content_type=None)
                        qty_map = {
                            str(item["locationId"]): item.get("inStockQuantity", 0)
                            for item in data.get("items", [])
                        }
                    else:
                        qty_map = {}
            except Exception as e:
                log(f"Error fetching store stock: {e}")
                qty_map = {}
        
        store_lines = []
        for store_id, city, address in user_stores:
            addr_line = ", ".join(filter(None, [address, city]))
            qty = qty_map.get(store_id, "(error fetching stock)")
            line = f"<b>{qty}</b> in stock — {store_id} — {addr_line}"
            store_lines.append(line)
        
        if store_lines:
            msg += "\n\nStock at your stores:\n" + "\n".join(store_lines)
    
    if info.get("Thumbnail"):
        await update.message.reply_photo(
            photo=info["Thumbnail"],
            caption=msg,
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(msg, parse_mode="HTML")

# ============================================================
# GENERATE AND SEND DAILY REPORT
# ============================================================

# ============================================================
# GLOBAL REPORT GENERATION
# Add these functions to your bot after the database operations
# ============================================================

async def refresh_global_list(app=None, send_report_to_owner=True):
    """
    Build daily Excel report from global_products database (fresh data fetched),
    compare with yesterday, highlight changes, format, save, and send to all users.
    OPTIMIZED: Fetches products concurrently.
    """
    try:
        log("🔄 Running Daily Global Report...")
        log(f"App passed to function: {app is not None}")
        
        # Get all product IDs from database
        product_ids = get_all_global_product_ids()
        log(f"Found {len(product_ids)} products in database")

        if not product_ids:
            log("⚠️ No products in global_products; skipping report.")
            return
    except Exception as e:
        log(f"❌ Error at start of refresh_global_list: {e}")
        log(f"Traceback: {traceback.format_exc()}")
        return

    # Fetch fresh product info for each product ID CONCURRENTLY
    all_data = []
    log(f"Starting concurrent fetch for {len(product_ids)} products...")
    
    # Use asyncio to fetch multiple products at once
    import aiohttp
    
    async def fetch_product_async(pid, session):
        """Async wrapper for get_product_info"""
        headers = {"User-Agent": "Mozilla/5.0"}
        product_url = f"{PRODUCT_BASE_URL}/{pid}"
        stock_url = (
            f"{BASE_URL}/ccstore/v1/stockStatus"
            f"?actualStockStatus=true&expandStockDetails=true&products=repositoryId%3A{pid}&locationIds=9650"
        )
        
        try:
            # Fetch product data
            async with session.get(product_url, headers=headers, timeout=aiohttp.ClientTimeout(total=10)) as res:
                if res.status >= 400:
                    log(f"❌ Invalid product ID {pid} (HTTP {res.status})")
                    return None
                
                data = await res.json()
                
                parent_categories = data.get("parentCategories", [])
                parent_category = ", ".join(
                    cat.get("displayName", cat.get("repositoryId", "N/A"))
                    for cat in parent_categories
                ) if parent_categories else "N/A"
                
                active = data.get("active", "N/A")
                display_name = data.get("displayName", "Unknown")
                highly_allocated = data.get("b2c_highlyAllocatedProduct", "N/A")
                lottery_product = data.get("b2c_lotteryProduct", "N/A")
                repo_id = data.get("repositoryId", pid)
                order_limit = data.get("b2c_limitPerOrder", "N/A")
                route = data.get("route", "")
                product_full_url = f"{BASE_URL}{route}" if route else f"{BASE_URL}/product/{pid}"
                
                thumb_path = data.get("primarySmallImageURL")
                thumbnail_url = f"{BASE_URL}{thumb_path}" if thumb_path else None
                
                list_price = data.get("listPrice", "N/A")
                if isinstance(list_price, dict):
                    list_price = list_price.get("value", "N/A")
                
                # Fetch stock
                try:
                    async with session.get(stock_url, headers=headers, timeout=aiohttp.ClientTimeout(total=10)) as stock_res:
                        if stock_res.status < 400:
                            stock_data = await stock_res.json()
                            stock_info = stock_data.get("items", [])[0] if stock_data.get("items") else {}
                            quantity = stock_info.get("inStockQuantity", 0)
                            if not str(quantity).isdigit():
                                quantity = 0
                        else:
                            quantity = 0
                except Exception:
                    quantity = 0
                
                return {
                    "ProductID": str(repo_id),
                    "Name": display_name,
                    "Category": parent_category,
                    "Active": str(active),
                    "InStock": int(quantity),
                    "Allocated": highly_allocated,
                    "Lottery": lottery_product,
                    "Price": str(list_price),
                    "OdrLmt": order_limit,
                    "Thumbnail": thumbnail_url,
                    "product_full_url": product_full_url
                }
        except Exception as e:
            log(f"❌ Error fetching product {pid}: {e}")
            return None
    
    # Fetch all products concurrently in batches
    BATCH_SIZE = 20
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(product_ids), BATCH_SIZE):
            batch = product_ids[i:i+BATCH_SIZE]
            log(f"Fetching batch {i//BATCH_SIZE + 1}/{(len(product_ids)-1)//BATCH_SIZE + 1} ({len(batch)} products)...")
            
            # Fetch all in batch concurrently
            tasks = [fetch_product_async(pid, session) for pid in batch]
            results = await asyncio.gather(*tasks)
            
            # Process results
            for j, product_info in enumerate(results):
                if product_info:
                    log(f"[{i+j+1}/{len(product_ids)}] Fetched: {batch[j]} | {product_info.get('Name')}")
                    all_data.append(product_info)
                    # Update database with fresh data
                    add_to_global_products(product_info)
                else:
                    # Use placeholder for failed fetches
                    pid = batch[j]
                    product_info = {
                        "ProductID": str(pid),
                        "Name": "Unknown",
                        "Category": "N/A",
                        "Active": "N/A",
                        "InStock": 0,
                        "Price": 0,
                        "Allocated": "N/A",
                        "Lottery": "N/A",
                        "OdrLmt": "N/A"
                    }
                    all_data.append(product_info)
            
            # Small delay between batches
            if i + BATCH_SIZE < len(product_ids):
                await asyncio.sleep(0.5)
    
    log(f"✅ Fetched data for {len(all_data)} products")
    
    if not all_data:
        log("❌ No data collected, cannot create report")
        return

    # Build DataFrame
    try:
        log("Building DataFrame...")
        df = pd.DataFrame(all_data)
        df.columns = df.columns.str.strip()
        log(f"DataFrame created with {len(df)} rows and columns: {list(df.columns)}")
    except Exception as e:
        log(f"❌ Error building DataFrame: {e}")
        log(f"Traceback: {traceback.format_exc()}")
        return

    # Remove unwanted columns if present
    for col in ["Modified By", "LastModified", "Thumbnail", "product_full_url"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    # Normalize ProductID column name
    if "Id" in df.columns:
        df.rename(columns={"Id": "ProductID"}, inplace=True)
    if "ProductID" in df.columns:
        df["ProductID"] = df["ProductID"].astype(str)
    else:
        df["ProductID"] = df.index.astype(str)

    # Ensure InStock numeric
    if "InStock" in df.columns:
        df["InStock"] = pd.to_numeric(df["InStock"], errors="coerce").fillna(0).astype(int)
    else:
        df["InStock"] = 0
    
    # Sort by Name (alphabetically, case-insensitive)
    if "Name" in df.columns:
        df = df.sort_values(by="Name", key=lambda x: x.str.lower())
        log("✅ Sorted by product name")
    
    # Reset index after sorting
    df = df.reset_index(drop=True)

    # Save today's raw DataFrame to Excel (initial)
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        filename = f"{REPORT_PREFIX}{today_str}.xlsx"
        log(f"Attempting to save Excel file: {filename}")
        log(f"Current working directory: {os.getcwd()}")
        
        df.to_excel(filename, index=False)
        
        if os.path.exists(filename):
            file_size = os.path.getsize(filename)
            log(f"✅ Saved daily report: {filename} ({file_size} bytes)")
        else:
            log(f"❌ File not created: {filename}")
            return
    except Exception as e:
        log(f"❌ Error saving Excel file: {e}")
        log(f"Traceback: {traceback.format_exc()}")
        return

    # Compare with yesterday's report and highlight changes
    yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    yesterday_filename = f"{REPORT_PREFIX}{yesterday_str}.xlsx"

    if os.path.exists(yesterday_filename):
        log(f"🔎 Comparing with yesterday's report: {yesterday_filename}")
        df_yesterday = pd.read_excel(yesterday_filename)
        df_yesterday.columns = df_yesterday.columns.str.strip()
        df.columns = df.columns.str.strip()

        if "Id" in df_yesterday.columns:
            df_yesterday.rename(columns={"Id": "ProductID"}, inplace=True)

        df["ProductID"] = df["ProductID"].astype(str)
        if "ProductID" in df_yesterday.columns:
            df_yesterday["ProductID"] = df_yesterday["ProductID"].astype(str)

        for col in ["InStock", "OdrLmt", "Price"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            if col in df_yesterday.columns:
                df_yesterday[col] = pd.to_numeric(df_yesterday[col], errors='coerce')

        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            pid_cell = ws[f"A{row}"].value
            pid = str(pid_cell).strip() if pid_cell else ""
            if pid and "ProductID" in df_yesterday.columns and pid in df_yesterday["ProductID"].values:
                for col_name in df.columns:
                    if col_name not in df_yesterday.columns:
                        continue
                    col_idx = df.columns.get_loc(col_name) + 1
                    cell = ws.cell(row=row, column=col_idx)

                    old_val_series = df_yesterday.loc[df_yesterday["ProductID"] == pid, col_name]
                    if old_val_series.empty:
                        continue
                    old_val = old_val_series.values[0]
                    new_val = cell.value

                    old_norm = normalize_for_comparison(old_val, col_name)
                    new_norm = normalize_for_comparison(new_val, col_name)

                    if old_norm is None and new_norm is None:
                        continue
                    if old_norm != new_norm:
                        log(f"Change found for ProductID {pid}, col {col_name}: {old_norm} -> {new_norm}")
                        cell.fill = highlight

        wb.save(filename)
        log("✅ Changes highlighted in today's report")
    else:
        log("⚠️ No report from yesterday found, skipping comparison.")

    # Load workbook for formatting
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Find InStock and Price columns
    instock_col_idx = None
    price_col_idx = None
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value).strip().lower()
        if header == "instock":
            instock_col_idx = col
        elif header == "price":
            price_col_idx = col

    # Remove existing "Total" if present
    if ws.cell(row=ws.max_row, column=1).value == "Total":
        ws.delete_rows(ws.max_row)

    last_data_row = ws.max_row

    if last_data_row < 2:
        log("⚠️ No data rows found to sum for InStock.")
    elif instock_col_idx is not None and price_col_idx is not None:
        instock_values = []
        total_inventory_value = 0
        
        for row in range(2, last_data_row + 1):
            instock_cell = ws.cell(row=row, column=instock_col_idx)
            price_cell = ws.cell(row=row, column=price_col_idx)

            instock_val = to_number(instock_cell.value)
            price_val = to_number(price_cell.value)

            instock_cell.value = instock_val
            instock_cell.alignment = Alignment(horizontal="right")
            instock_cell.number_format = '#,##0'

            price_cell.value = price_val
            price_cell.alignment = Alignment(horizontal="right")
            price_cell.number_format = '#,##0.00'

            instock_values.append(instock_val)
            total_inventory_value += instock_val * price_val

        total_instock = sum(instock_values)
        total_row = ws.max_row + 1

        total_label = ws.cell(row=total_row, column=1)
        total_label.value = "Total"
        total_label.font = Font(bold=True)

        total_instock_cell = ws.cell(row=total_row, column=instock_col_idx)
        total_instock_cell.value = total_instock
        total_instock_cell.alignment = Alignment(horizontal="right")
        total_instock_cell.number_format = '#,##0'
        total_instock_cell.font = Font(bold=True)

        total_price_cell = ws.cell(row=total_row, column=price_col_idx)
        total_price_cell.value = total_inventory_value
        total_price_cell.alignment = Alignment(horizontal="right")
        total_price_cell.number_format = '#,##0.00'
        total_price_cell.font = Font(bold=True)

    else:
        log("⚠️ 'InStock' or 'Price' column not found, skipping numeric formatting.")

    # Smart auto-fit with custom width limits
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        adjusted_width = (max_length * 1.1) + 2
        header = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()

        # Apply per-column caps
        if header == "name":
            adjusted_width = min(adjusted_width, 55)
        elif header == "category":
            adjusted_width = min(adjusted_width, 18)
        elif header == "price":
            adjusted_width = min(adjusted_width, 12)
        else:
            if adjusted_width < 8:
                adjusted_width = 8
            elif adjusted_width > 40:
                adjusted_width = 40

        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save final workbook
    wb.save(filename)
    log(f"\n✅ Excel file saved, formatted: {filename}")

    # Send the report to all users
    if app:
        log("App object available, attempting to send reports...")
        try:
            await send_today_report_to_all_users(app, filename)
        except Exception as e:
            log(f"❌ Error sending reports: {e}")
            log(f"Traceback: {traceback.format_exc()}")
    else:
        log("⚠️ No app object provided, cannot send reports to users")
    
    log("✅ refresh_global_list completed")


async def send_today_report_to_all_users(app, filename):
    """Send the daily report to all users who have watchlists."""
    try:
        log(f"Starting send_today_report_to_all_users with file: {filename}")
        
        # Verify file exists
        if not os.path.exists(filename):
            log(f"❌ File does not exist: {filename}")
            return
        
        bot = app.bot
        log(f"Bot object: {bot is not None}")
        
        # Get all users
        user_ids = get_all_users()
        log(f"Found {len(user_ids)} users to send report to: {user_ids}")
        
        if not user_ids:
            log("⚠️ No users to send report to")
            return
        
        log(f"📤 Sending report to {len(user_ids)} users...")
        
        # Send to each user
        success_count = 0
        fail_count = 0
        
        for user_id in user_ids:
            try:
                log(f"Attempting to send to user {user_id}...")
                with open(filename, 'rb') as f:
                    await bot.send_document(
                        chat_id=int(user_id),
                        document=f,
                        caption="🥃 Your daily FWGS product report! Changes from yesterday are highlighted in yellow.",
                        filename=os.path.basename(filename)
                    )
                log(f"✅ Sent report to user {user_id}")
                success_count += 1
            except Exception as e:
                log(f"❌ Failed to send report to user {user_id}: {e}")
                fail_count += 1
        
        log(f"✅ Report distribution complete: {success_count} succeeded, {fail_count} failed")
    
    except Exception as e:
        log(f"❌ Error in send_today_report_to_all_users: {e}")
        log(f"Traceback: {traceback.format_exc()}")
    

def get_all_global_product_ids():
    """Get all product IDs from global_products table."""
    query = "SELECT product_id FROM global_products;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        return [row[0] for row in results]
    except Exception as e:
        log(f"❌ Error getting product IDs: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

def get_all_users():
    """Get all user IDs for sending reports."""
    query = "SELECT DISTINCT user_id FROM watchlist;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        return [row[0] for row in results]
    except Exception as e:
        log(f"❌ Error getting users: {e}")
        return []
    finally:
        if conn:
            return_db(conn)
            
async def send_global_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler for /lastreport command - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ Subscribe to receive daily reports: /subscribe"
        )
        return
    
    today_str = datetime.now().strftime("%Y%m%d")
    filename = f"{REPORT_PREFIX}{today_str}.xlsx"
    
    if not os.path.exists(filename):
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
        filename = f"{REPORT_PREFIX}{yesterday_str}.xlsx"
    
    if not os.path.exists(filename):
        await update.message.reply_text("⚠️ No recent report found. The next report will be generated at 9 AM.")
        return
    
    try:
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                caption="🥃 Here's the most recent FWGS product report!",
                filename=os.path.basename(filename)
            )
        log(f"✅ Sent report to user {update.effective_user.id}")
    except Exception as e:
        log(f"❌ Error sending report: {e}")
        await update.message.reply_text("⚠️ Error sending report. Please try again later.")

async def test_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test command to manually trigger report generation."""
    await update.message.reply_text("🔄 Starting report generation test...")
    
    try:
        # Check if pandas and openpyxl are available
        try:
            import pandas as pd
            import openpyxl
            await update.message.reply_text("✅ pandas and openpyxl are installed")
        except ImportError as e:
            await update.message.reply_text(f"❌ Missing library: {e}")
            return
        
        # Check database
        product_ids = get_all_global_product_ids()
        await update.message.reply_text(f"✅ Found {len(product_ids)} products in database")
        
        if not product_ids:
            await update.message.reply_text("⚠️ No products in global list. Add some with /add <ProductID>")
            return
        
        # Check users
        users = get_all_users()
        await update.message.reply_text(f"✅ Found {len(users)} users with watchlists")
        
        # Trigger report generation
        await update.message.reply_text("🔄 Generating report...")
        await refresh_global_list(context.application, send_report_to_owner=True)
        
        await update.message.reply_text("✅ Report generation completed! Check logs for details.")
        
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}\n\nCheck logs for full traceback.")
        log(f"❌ Error in test_report: {e}")
        log(f"Traceback: {traceback.format_exc()}")


# ============================================================
# MONITORING HELPER FUNCTIONS - ACTIVE and CATEGORY
# ============================================================

async def get_active_only(pid, session):
    """
    Lightweight: fetch ONLY the active flag for a product.
    Returns: (pid, True/False) or (pid, None) on error.
    Note: Session must be provided (don't create/close inside)
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    url = f"{PRODUCT_BASE_URL}/{pid}?fields=active"
    
    try:
        async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=5)) as resp:
            if resp.status != 200:
                return (pid, None)
            data = await resp.json()
            return (pid, bool(data.get("active")) if "active" in data else None)
    except Exception as e:
        log(f"⚠️ Error fetching active for {pid}: {e}")
        return (pid, None)


async def get_category_only(pid, session):
    """
    Lightweight: fetch ONLY the parentCategories list.
    Returns: (pid, [category_ids]) or (pid, None) on error.
    """
    headers = {"User-Agent": "Mozilla/5.0"}
    url = f"{PRODUCT_BASE_URL}/{pid}?fields=parentCategories"
    
    try:
        async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=5)) as resp:
            if resp.status != 200:
                return (pid, None)
            data = await resp.json()
            parent_categories = data.get("parentCategories", [])
            categories = [c.get("repositoryId", "") for c in parent_categories if "repositoryId" in c]
            return (pid, categories)
    except Exception as e:
        log(f"⚠️ Error fetching categories for {pid}: {e}")
        return (pid, None)

def get_product_active_state(product_id):
    """Get the last known active state from product_cache."""
    query = "SELECT last_qty FROM product_cache WHERE product_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        result = cur.fetchone()
        cur.close()
        return bool(result[0]) if result else None
    except Exception as e:
        log(f"❌ Error getting product state: {e}")
        return None
    finally:
        if conn:
            return_db(conn)

def get_product_active_state(product_id):
    """Get the last known active state from product_cache."""
    query = "SELECT last_qty FROM product_cache WHERE product_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        result = cur.fetchone()
        cur.close()
        return bool(result[0]) if result else None
    except Exception as e:
        log(f"❌ Error getting product state: {e}")
        return None
    finally:
        if conn:
            return_db(conn)

def set_product_categories(product_id, categories):
    """Store product categories in product_cache."""
    query = """
        INSERT INTO product_cache (product_id, category, updated)
        VALUES (%s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id) DO UPDATE SET
            category = EXCLUDED.category,
            updated = CURRENT_TIMESTAMP;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        categories_str = ",".join([str(c).lower() for c in categories])
        cur.execute(query, (product_id, categories_str))
        conn.commit()
        cur.close()
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error setting product categories: {e}")
    finally:
        if conn:
            return_db(conn)

def update_store_quantity(product_id, store_id, quantity):
    """Update the quantity for a product at a specific store."""
    query = """
        INSERT INTO product_quantity_cache (product_id, store_id, last_qty, last_checked)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id, store_id) DO UPDATE SET
            last_qty = EXCLUDED.last_qty,
            last_checked = CURRENT_TIMESTAMP;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id, store_id, int(quantity)))
        conn.commit()
        cur.close()
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error updating store quantity: {e}")
    finally:
        if conn:
            return_db(conn)

def users_watching_product(product_id):
    """Get list of user_ids watching a specific product."""
    query = "SELECT DISTINCT user_id FROM watchlist WHERE product_id = %s;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        results = cur.fetchall()
        cur.close()
        return [row[0] for row in results]
    except Exception as e:
        log(f"❌ Error getting watchers: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

def get_last_store_quantity(product_id, store_id):
    """Get the last known quantity for a product at a specific store."""
    query = """
        SELECT last_qty 
        FROM product_quantity_cache 
        WHERE product_id = %s AND store_id = %s 
        LIMIT 1;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id, store_id))
        result = cur.fetchone()
        cur.close()
        return int(result[0]) if result else 0
    except Exception as e:
        log(f"❌ Error getting last quantity: {e}")
        return 0
    finally:
        if conn:
            return_db(conn)

def get_product_categories(product_id):
    """Get the last known categories from product_cache."""
    query = "SELECT category FROM product_cache WHERE product_id = %s LIMIT 1;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        result = cur.fetchone()
        cur.close()
        if result and result[0]:
            return result[0].split(",") if result[0] else []
        return []
    except Exception as e:
        log(f"❌ Error getting product categories: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

def get_product_active_states_batch(product_ids):
    """Get active states for multiple products at once."""
    if not product_ids:
        return {}
    
    query = """
        SELECT product_id, last_qty 
        FROM product_cache 
        WHERE product_id = ANY(%s);
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_ids,))
        results = cur.fetchall()
        cur.close()
        # Return dict: {product_id: bool(active)}
        return {pid: bool(active) for pid, active in results}
    except Exception as e:
        log(f"❌ Error getting batch product states: {e}")
        return {}
    finally:
        if conn:
            return_db(conn)


def set_product_active_states_batch(updates):
    """Store active states for multiple products at once.
    
    Args:
        updates: dict of {product_id: is_active}
    """
    if not updates:
        return
    
    query = """
        INSERT INTO product_cache (product_id, last_qty, updated)
        VALUES (%s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id) DO UPDATE SET
            last_qty = EXCLUDED.last_qty,
            updated = CURRENT_TIMESTAMP;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Prepare batch data
        data = [(pid, 1 if active else 0) for pid, active in updates.items()]
        cur.executemany(query, data)
        conn.commit()
        cur.close()
        log(f"✅ Updated {len(updates)} product states in batch")
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error setting batch product states: {e}")
    finally:
        if conn:
            return_db(conn)


def get_product_categories_batch(product_ids):
    """Get categories for multiple products at once."""
    if not product_ids:
        return {}
    
    query = """
        SELECT product_id, category 
        FROM product_cache 
        WHERE product_id = ANY(%s);
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_ids,))
        results = cur.fetchall()
        cur.close()
        # Return dict: {product_id: [categories]}
        return {
            pid: cats.split(",") if cats else [] 
            for pid, cats in results
        }
    except Exception as e:
        log(f"❌ Error getting batch categories: {e}")
        return {}
    finally:
        if conn:
            return_db(conn)


def set_product_categories_batch(updates):
    """Store categories for multiple products at once.
    
    Args:
        updates: dict of {product_id: [categories]}
    """
    if not updates:
        return
    
    query = """
        INSERT INTO product_cache (product_id, category, updated)
        VALUES (%s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id) DO UPDATE SET
            category = EXCLUDED.category,
            updated = CURRENT_TIMESTAMP;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Prepare batch data
        data = [
            (pid, ",".join([str(c).lower() for c in cats]))
            for pid, cats in updates.items()
        ]
        cur.executemany(query, data)
        conn.commit()
        cur.close()
        log(f"✅ Updated {len(updates)} product categories in batch")
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error setting batch categories: {e}")
    finally:
        if conn:
            return_db(conn)


# ============================================================
# ACTIVE AND CATEGORY MONITORING
# ============================================================
async def active_monitor(context: ContextTypes.DEFAULT_TYPE):
    """Monitor products for active status changes - OPTIMIZED VERSION."""
    try:
        import aiohttp  # Import here too just to be safe
        
        bot = context.bot
        
        # Get all global products from database
        global_cache = get_all_global_products()
        
        if not global_cache:
            log("⚠️ No products in global list to monitor")
            return
        
        product_ids = list(global_cache.keys())
        log(f"⏰ Active monitor checking {len(product_ids)} products...")
        
        # Get previous states in one batch query
        prev_states = get_product_active_states_batch(product_ids)
        
        # Fetch current states concurrently (batches of 20 at a time to avoid overwhelming API)
        BATCH_SIZE = 20
        current_states = {}
        
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(product_ids), BATCH_SIZE):
                batch = product_ids[i:i+BATCH_SIZE]
                
                # Fetch all in batch concurrently
                tasks = [get_active_only(pid, session) for pid in batch]
                results = await asyncio.gather(*tasks)
                
                # Store results
                for pid, active_now in results:
                    if active_now is not None:
                        current_states[pid] = active_now
                
                # Small delay between batches to be nice to the API
                if i + BATCH_SIZE < len(product_ids):
                    await asyncio.sleep(0.5)
        
        log(f"✅ Fetched active states for {len(current_states)} products")
        
        # Check for changes and send alerts
        changes = {}
        alerts_to_send = []
        
        for pid, active_now in current_states.items():
            prev_active = prev_states.get(pid)
            
            if prev_active is None:
                # First time seeing this product, just store state
                changes[pid] = active_now
                continue
            
            # Check if status changed
            if prev_active != active_now:
                changes[pid] = active_now
                
                info = global_cache.get(pid, {})
                name = info.get("Name", "Unknown")
                url = info.get("product_full_url", "")
                
                # Build alert message
                if active_now:
                    msg = f"🔥 <a href='{url}'>{name}</a> is ACTIVE!"
                else:
                    msg = f"⚠️ {name} is INACTIVE."
                
                # Get all users watching this product
                watchers = users_watching_product(pid)
                
                # Queue alerts
                for user_id in watchers:
                    alerts_to_send.append((user_id, msg))
        
        # Update all changed states in one batch
        if changes:
            set_product_active_states_batch(changes)
            log(f"✅ Detected {len(changes)} active state changes")
        
        # Send all alerts
        if alerts_to_send:
            log(f"📤 Sending {len(alerts_to_send)} alerts...")
            for user_id, msg in alerts_to_send:
                try:
                    await bot.send_message(
                        chat_id=int(user_id), 
                        text=msg, 
                        parse_mode="HTML"
                    )
                except Exception as e:
                    log(f"❌ Failed to send alert to {user_id}: {e}")
        
        log("✅ Active monitor completed")
        
    except Exception as e:
        log(f"❌ Error in active_monitor: {e}")
        log(f"Traceback: {traceback.format_exc()}")


async def category_monitor(context: ContextTypes.DEFAULT_TYPE):
    """Monitor products for category changes - OPTIMIZED VERSION."""
    try:
        import aiohttp  # Import here too
        
        bot = context.bot
        
        # Get all global products from database
        global_cache = get_all_global_products()
        
        if not global_cache:
            log("⚠️ No products in global list to monitor")
            return
        
        product_ids = list(global_cache.keys())
        log(f"⏰ Category monitor checking {len(product_ids)} products...")
        
        # Get previous categories in one batch query
        prev_categories = get_product_categories_batch(product_ids)
        
        # Fetch current categories concurrently
        BATCH_SIZE = 20
        current_categories = {}
        
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(product_ids), BATCH_SIZE):
                batch = product_ids[i:i+BATCH_SIZE]
                
                # Fetch all in batch concurrently
                tasks = [get_category_only(pid, session) for pid in batch]
                results = await asyncio.gather(*tasks)
                
                # Store results
                for pid, categories in results:
                    if categories is not None:
                        current_categories[pid] = [str(c).lower() for c in categories]
                
                # Small delay between batches
                if i + BATCH_SIZE < len(product_ids):
                    await asyncio.sleep(0.5)
        
        log(f"✅ Fetched categories for {len(current_categories)} products")
        
        # Check for new whiskey-release categories
        updates = {}
        alerts_to_send = []
        
        for pid, cats_now in current_categories.items():
            prev_cats = prev_categories.get(pid, [])
            
            # Update storage
            updates[pid] = cats_now
            
            # Detect new whiskey-release category
            if "whiskey-release" in cats_now and "whiskey-release" not in prev_cats:
                info = global_cache.get(pid, {})
                name = info.get("Name", pid)
                msg = f"📣 Whiskey-release added for {name}!"
                
                # Get all users watching this product
                watchers = users_watching_product(pid)
                
                # Queue alerts
                for user_id in watchers:
                    alerts_to_send.append((user_id, msg))
        
        # Update all categories in one batch
        if updates:
            set_product_categories_batch(updates)
        
        # Send all alerts
        if alerts_to_send:
            log(f"📤 Sending {len(alerts_to_send)} category alerts...")
            for user_id, msg in alerts_to_send:
                try:
                    await bot.send_message(chat_id=int(user_id), text=msg)
                except Exception as e:
                    log(f"❌ Failed to send alert to {user_id}: {e}")
        
        log("✅ Category monitor completed")
        
    except Exception as e:
        log(f"❌ Error in category_monitor: {e}")
        log(f"Traceback: {traceback.format_exc()}")


# ============================================================
# HELPER FUNCTIONS FOR STOCK MONITORING
# ============================================================
def get_last_store_quantity(product_id, store_id):
    """Get the last known quantity for a product at a specific store."""
    query = """
        SELECT last_qty 
        FROM product_quantity_cache 
        WHERE product_id = %s AND store_id = %s 
        LIMIT 1;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id, store_id))
        result = cur.fetchone()
        cur.close()
        return int(result[0]) if result else 0
    except Exception as e:
        log(f"❌ Error getting last quantity: {e}")
        return 0
    finally:
        if conn:
            return_db(conn)


def update_store_quantity(product_id, store_id, quantity):
    """Update the quantity for a product at a specific store."""
    query = """
        INSERT INTO product_quantity_cache (product_id, store_id, last_qty, last_checked)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id, store_id) DO UPDATE SET
            last_qty = EXCLUDED.last_qty,
            last_checked = CURRENT_TIMESTAMP;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id, store_id, int(quantity)))
        conn.commit()
        cur.close()
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error updating store quantity: {e}")
    finally:
        if conn:
            return_db(conn)

def users_watching_product(product_id):
    """Get list of user_ids watching a specific product."""
    query = "SELECT DISTINCT user_id FROM watchlist WHERE product_id = %s;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (product_id,))
        results = cur.fetchall()
        cur.close()
        return [row[0] for row in results]  # Return list of user_ids
    except Exception as e:
        log(f"❌ Error getting watchers: {e}")
        return []
    finally:
        if conn:
            return_db(conn)

def get_all_global_products():
    """Get all products from global_products table."""
    query = "SELECT product_id, name, product_full_url FROM global_products;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        # Return as dict: {product_id: {"Name": name, "product_full_url": url}, ...}
        return {pid: {"Name": name, "product_full_url": url} for pid, name, url in results}
    except Exception as e:
        log(f"❌ Error getting global products: {e}")
        return {}
    finally:
        if conn:
            return_db(conn)

async def inventory_refresh_job(context: ContextTypes.DEFAULT_TYPE):
    """
    Scheduled job: scans users' watchlists and stores, sends alerts on stock increases.
    Runs every 30 minutes during business hours.
    """
    import aiohttp
    from datetime import datetime
    
    bot = context.bot
    
    # Check business hours
    now = datetime.now().time()
    if not (BUSINESS_START <= now <= BUSINESS_END):
        log(f"Outside business hours ({BUSINESS_START}-{BUSINESS_END}). Skipping inventory check.")
        return
    
    log("⏰ Inventory refresh job starting...")
    
    # Get all users who have both watchlist items and stores
    query = """
        SELECT DISTINCT w.user_id, w.user_name
        FROM watchlist w
        INNER JOIN stores s ON w.user_id = s.user_id
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        users = cur.fetchall()
        cur.close()
    except Exception as e:
        log(f"❌ Error getting users for inventory check: {e}")
        return
    finally:
        if conn:
            return_db(conn)
    
    if not users:
        log("No users with both watchlist and stores found.")
        return
    
    log(f"Checking inventory for {len(users)} users...")
    
    async with aiohttp.ClientSession() as session:
        for user_id, user_name in users:
            # Get user's watchlist
            watchlist = get_user_watchlist(user_id)
            if not watchlist:
                continue
            
            # Get user's stores
            stores = get_user_stores(user_id)
            if not stores:
                continue
            
            # Build location IDs string for API
            location_ids_str = ",".join([store_id for store_id, _, _ in stores])
            headers = {"User-Agent": "Mozilla/5.0"}
            
            # Check each product in watchlist
            for product_id, product_name in watchlist:
                # Build API URL
                url = (
                    f"{BASE_URL}/ccstore/v1/stockStatus"
                    f"?actualStockStatus=true"
                    f"&expandStockDetails=true"
                    f"&products=repositoryId:{product_id}"
                    f"&locationIds={location_ids_str}"
                )
                
                # Fetch current stock levels
                try:
                    async with session.get(
                        url, 
                        timeout=aiohttp.ClientTimeout(total=10), 
                        headers=headers
                    ) as resp:
                        if resp.status != 200:
                            log(f"Failed to fetch stock for {product_id}")
                            continue
                        
                        data = await resp.json(content_type=None)
                        qty_map = {
                            str(item["locationId"]): item.get("inStockQuantity", 0)
                            for item in data.get("items", [])
                        }
                
                except Exception as e:
                    log(f"❌ Exception fetching stock for {product_id}: {e}")
                    continue
                
                # Check each store for stock increases
                for store_id, city, address in stores:
                    current_qty = int(qty_map.get(store_id, 0))
                    
                    # Get last known quantity from database
                    last_qty = get_last_store_quantity(product_id, store_id)
                    
                    # Alert on stock increase
                    if current_qty > last_qty:
                        try:
                            text = (
                                "🔔 <b>Stock Added!</b>\n\n"
                                f"<b>Product:</b> {product_id} - {product_name}\n"
                                f"<b>Store:</b> {store_id} - {city} - {address}\n"
                                f"<b>Quantity:</b> <b>{last_qty} ➜ {current_qty}</b>"
                            )
                            await bot.send_message(
                                chat_id=int(user_id),
                                text=text,
                                parse_mode="HTML"
                            )
                            log(f"✅ Sent stock alert to {user_id} for {product_id} at store {store_id}")
                        
                        except Exception as e:
                            log(f"❌ Failed to send stock alert to {user_id}: {e}")
                    
                    # Update quantity in database (always, even if no change)
                    update_store_quantity(product_id, store_id, current_qty)
    
    log("✅ Inventory refresh job completed")

# ============================================================
# DATABASE OPERATIONS - FWGS STORES
# ============================================================

def load_fwgs_stores_from_json(json_file_path):
    """One-time import: Load all FWGS stores from JSON into database."""
    import json
    
    if not os.path.exists(json_file_path):
        log(f"❌ JSON file not found: {json_file_path}")
        return 0
    
    with open(json_file_path, "r") as f:
        all_stores = json.load(f)
    
    query = """
        INSERT INTO fwgs_stores (store_id, city, address1)
        VALUES (%s, %s, %s)
        ON CONFLICT (store_id) DO UPDATE SET
            city = EXCLUDED.city,
            address1 = EXCLUDED.address1,
            last_updated = CURRENT_TIMESTAMP;
    """
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Prepare data for batch insert
        data = []
        for store_id, store_data in all_stores.items():
            city = store_data.get("city", "Unknown")
            address = store_data.get("address1", "Unknown")
            data.append((store_id, city, address))
        
        # Batch insert
        cur.executemany(query, data)
        conn.commit()
        count = len(data)
        cur.close()
        log(f"✅ Loaded {count} stores into database")
        return count
        
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error loading stores: {e}")
        return 0
    finally:
        if conn:
            return_db(conn)


def get_all_fwgs_stores():
    """Get all FWGS stores from database.
    Returns: dict of {store_id: {"city": ..., "address1": ...}}
    """
    query = "SELECT store_id, city, address1 FROM fwgs_stores ORDER BY store_id;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        
        # Return as dict for easy lookup
        return {
            store_id: {"city": city, "address1": address1}
            for store_id, city, address1 in results
        }
    except Exception as e:
        log(f"❌ Error getting FWGS stores: {e}")
        return {}
    finally:
        if conn:
            return_db(conn)


def get_fwgs_stores_list():
    """Get all FWGS stores as list of tuples for chunked processing.
    Returns: [(store_id, city, address1), ...]
    """
    query = "SELECT store_id, city, address1 FROM fwgs_stores ORDER BY store_id;"
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        return results
    except Exception as e:
        log(f"❌ Error getting FWGS stores list: {e}")
        return []
    finally:
        if conn:
            return_db(conn)


# ============================================================
# STATE STOCK HANDLER
# ============================================================

async def statestock_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Check stock at ALL Pennsylvania FWGS stores for a product - REQUIRES SUBSCRIPTION."""
    user_id = str(update.effective_user.id)
    
    # Check access
    has_access, reason = check_access(user_id)
    if not has_access:
        await update.message.reply_text(
            "⭐ <b>Premium Feature</b>\n\n"
            "Subscribe to check statewide inventory:\n/subscribe",
            parse_mode="HTML"
        )
        return
    
    import aiohttp
    
    if not context.args:
        await update.message.reply_text("Usage: /statestock <product_id>")
        return
    
    pid = context.args[0].strip()
    await update.message.reply_text(f"🔎 Fetching statewide stock info for {pid}...")
    
    store_items = get_fwgs_stores_list()
    
    if not store_items:
        await update.message.reply_text("⚠️ No stores found in database. Contact admin.")
        return
    
    log(f"Checking stock at {len(store_items)} stores for product {pid}")
    
    headers = {"User-Agent": "Mozilla/5.0"}
    stock_results = []
    
    CHUNK_SIZE = 50
    
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(store_items), CHUNK_SIZE):
            chunk = store_items[i:i + CHUNK_SIZE]
            location_ids_str = ",".join([store_id for store_id, _, _ in chunk])
            url = FWGS_STOCKSTATUS_URL.format(pid=pid, store=location_ids_str)
            
            try:
                async with session.get(
                    url, 
                    headers=headers, 
                    timeout=aiohttp.ClientTimeout(total=10)
                ) as resp:
                    if resp.status == 200:
                        data = await resp.json(content_type=None)
                        store_lookup = {
                            store_id: (city, address1) 
                            for store_id, city, address1 in chunk
                        }
                        
                        for item in data.get("items", []):
                            loc_id = str(item["locationId"])
                            qty = item.get("inStockQuantity", 0)
                            
                            if qty > 0:
                                city, address = store_lookup.get(loc_id, ("Unknown", "Unknown"))
                                stock_results.append(f"<b>{qty}</b> in stock — {loc_id} — {city}, {address}")
                    
                    elif resp.status != 204:
                        log(f"⚠️ Failed chunk {i//CHUNK_SIZE + 1} (status: {resp.status})")
                        
            except Exception as e:
                log(f"❌ Error fetching stock chunk {i//CHUNK_SIZE + 1}: {e}")
                continue
            
            if i + CHUNK_SIZE < len(store_items):
                await asyncio.sleep(0.3)
    
    stock_results.sort(key=lambda x: int(x.split("<b>")[1].split("</b>")[0]), reverse=True)
    
    if stock_results:
        MAX_CHARS = 4000
        message = f"📦 <b>Statewide Stock for {pid}</b>\n\n"
        
        for line in stock_results:
            if len(message) + len(line) + 1 > MAX_CHARS:
                await update.message.reply_text(message, parse_mode="HTML")
                message = ""
            message += line + "\n"
        
        if message:
            total_stores = len(stock_results)
            total_qty = sum(int(line.split("<b>")[1].split("</b>")[0]) for line in stock_results)
            message += f"\n<b>Total: {total_qty} units at {total_stores} stores</b>"
            await update.message.reply_text(message, parse_mode="HTML")
        
        log(f"✅ Found stock at {len(stock_results)} stores")
    else:
        await update.message.reply_text(f"📭 No stock found at any store for product {pid}.")

# ============================================================
# SUBSCRIPTION OPERATIONS - USER MANAGEMENT
# ============================================================
def create_or_update_user(user_id, full_name):
    """Create user or update their info. New users get free trial."""
    query = """
        INSERT INTO users (user_id, full_name, is_subscribed, subscription_expiry, joined)
        VALUES (%s, %s, TRUE, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (user_id) DO UPDATE SET
            full_name = EXCLUDED.full_name;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # New users get 2-week trial
        trial_expiry = datetime.now() + timedelta(days=TRIAL_DAYS)
        
        cur.execute(query, (user_id, full_name, trial_expiry))
        conn.commit()
        cur.close()
        log(f"✅ Created user {user_id} with trial until {trial_expiry}")
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error creating user: {e}")
    finally:
        if conn:
            return_db(conn)


def get_user_subscription_status(user_id):
    """Get user's subscription details.
    Returns: dict with is_admin, is_subscribed, expiry, trial_active
    """
    query = """
        SELECT is_admin, is_subscribed, subscription_expiry 
        FROM users 
        WHERE user_id = %s;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id,))
        result = cur.fetchone()
        cur.close()
        
        if not result:
            return None
        
        is_admin, is_subscribed, expiry = result
        now = datetime.now()
        
        # Check if subscription is still valid
        is_active = is_subscribed and (expiry is None or expiry > now)
        
        # Check if in trial period (expiry within TRIAL_DAYS of account creation)
        query_joined = "SELECT joined FROM users WHERE user_id = %s;"
        cur = conn.cursor()
        cur.execute(query_joined, (user_id,))
        joined = cur.fetchone()
        cur.close()
        
        trial_active = False
        if joined and expiry:
            trial_end = joined[0] + timedelta(days=TRIAL_DAYS)
            trial_active = now < trial_end and is_active
        
        return {
            "is_admin": is_admin,
            "is_subscribed": is_subscribed,
            "expiry": expiry,
            "is_active": is_active,
            "trial_active": trial_active
        }
    except Exception as e:
        log(f"❌ Error getting subscription status: {e}")
        return None
    finally:
        if conn:
            return_db(conn)


def extend_subscription(user_id, days=SUBSCRIPTION_DURATION_DAYS):
    """Extend user's subscription by X days from current expiry (or now if expired)."""
    query = """
        UPDATE users 
        SET is_subscribed = TRUE,
            subscription_expiry = CASE 
                WHEN subscription_expiry > CURRENT_TIMESTAMP 
                THEN subscription_expiry + INTERVAL '%s days'
                ELSE CURRENT_TIMESTAMP + INTERVAL '%s days'
            END
        WHERE user_id = %s
        RETURNING subscription_expiry;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (days, days, user_id))
        result = cur.fetchone()
        conn.commit()
        new_expiry = result[0] if result else None
        cur.close()
        log(f"✅ Extended subscription for {user_id} until {new_expiry}")
        return new_expiry
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error extending subscription: {e}")
        return None
    finally:
        if conn:
            return_db(conn)


def set_admin_status(user_id, is_admin=True):
    """Grant or revoke admin status (admins don't need subscription)."""
    query = """
        INSERT INTO users (user_id, is_admin, is_subscribed, subscription_expiry)
        VALUES (%s, %s, TRUE, NULL)
        ON CONFLICT (user_id) DO UPDATE SET
            is_admin = EXCLUDED.is_admin,
            is_subscribed = TRUE,
            subscription_expiry = NULL;
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query, (user_id, is_admin))
        conn.commit()
        cur.close()
        log(f"✅ Set admin status for {user_id}: {is_admin}")
        return True
    except Exception as e:
        if conn:
            conn.rollback()
        log(f"❌ Error setting admin status: {e}")
        return False
    finally:
        if conn:
            return_db(conn)


def check_access(user_id):
    """Check if user has access (admin OR active subscription).
    Returns: (has_access: bool, reason: str)
    """
    status = get_user_subscription_status(user_id)
    
    if not status:
        return False, "no_account"
    
    # Admins always have access
    if status["is_admin"]:
        return True, "admin"
    
    # Check if subscription is active
    if status["is_active"]:
        if status["trial_active"]:
            return True, "trial"
        return True, "subscribed"
    
    return False, "expired"

async def status_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show user's subscription status."""
    user_id = str(update.effective_user.id)
    status = get_user_subscription_status(user_id)
    
    if not status:
        await update.message.reply_text(
            "⚠️ No account found. Send /start to create one."
        )
        return
    
    if status["is_admin"]:
        msg = "👑 <b>Admin Account</b>\n\nYou have full access to all features."
    elif status["is_active"]:
        days_left = (status["expiry"] - datetime.now()).days
        if status["trial_active"]:
            msg = (
                f"🎁 <b>Free Trial Active</b>\n\n"
                f"Days remaining: {days_left}\n"
                f"Trial expires: {status['expiry'].strftime('%B %d, %Y')}\n\n"
                f"Subscribe now with /subscribe to continue after trial!"
            )
        else:
            msg = (
                f"✅ <b>Subscription Active</b>\n\n"
                f"Days remaining: {days_left}\n"
                f"Renews on: {status['expiry'].strftime('%B %d, %Y')}\n\n"
                f"Use /subscribe to extend your subscription."
            )
    else:
        msg = (
            "❌ <b>Subscription Expired</b>\n\n"
            "Your subscription has ended. Subscribe for $5/month to continue:\n"
            "/subscribe"
        )
    
    await update.message.reply_text(msg, parse_mode="HTML")

async def subscribe_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send subscription payment invoice."""
    user_id = str(update.effective_user.id)
    status = get_user_subscription_status(user_id)
    
    if status and status["is_admin"]:
        await update.message.reply_text("👑 You're an admin - no payment needed!")
        return
    
    # Send invoice
    await context.bot.send_invoice(
        chat_id=update.effective_chat.id,
        title="FWGS Bot Premium - 30 Days",
        description="30 days of premium access: product alerts, inventory tracking, and daily reports",
        payload=f"subscription_30d_{user_id}",
        provider_token="",  # Empty string for Telegram Stars
        currency="XTR",  # Telegram Stars currency code
        prices=[LabeledPrice("30 Day Access", SUBSCRIPTION_PRICE_STARS)],
        max_tip_amount=0,
        suggested_tip_amounts=[],
        start_parameter="subscribe",
        provider_data=None,
        #photo_url="https://www.finewineandgoodspirits.com/themes/custom/fwgs/logo.svg",
        #photo_size=100,
        #photo_width=200,
        #photo_height=200,
        need_name=False,
        need_phone_number=False,
        need_email=False,
        need_shipping_address=False,
        send_phone_number_to_provider=False,
        send_email_to_provider=False,
        is_flexible=False,
    )

async def precheckout_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle pre-checkout query (approve payment)."""
    query: PreCheckoutQuery = update.pre_checkout_query
    
    # Always approve (we handle everything after successful payment)
    await query.answer(ok=True)
    log(f"✅ Pre-checkout approved for user {query.from_user.id}")


async def successful_payment_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle successful payment - extend subscription."""
    payment = update.message.successful_payment
    user_id = str(update.effective_user.id)
    
    log(f"💰 Payment received from {user_id}: {payment.total_amount} Stars")
    
    # Extend subscription
    new_expiry = extend_subscription(user_id, SUBSCRIPTION_DURATION_DAYS)
    
    if new_expiry:
        await update.message.reply_text(
            f"🎉 <b>Payment Successful!</b>\n\n"
            f"Your subscription has been extended.\n"
            f"New expiry date: {new_expiry.strftime('%B %d, %Y')}\n\n"
            f"Thank you for subscribing! 🥃",
            parse_mode="HTML"
        )
        log(f"✅ Subscription extended for {user_id} until {new_expiry}")
    else:
        await update.message.reply_text(
            "⚠️ Payment received but there was an error updating your subscription. "
            "Please contact support."
        )
        
async def makeadmin_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Owner command to grant admin access. Usage: /makeadmin <user_id>"""
    # Only owner can use this
    if OWNER_CHAT_ID and str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ Owner only command.")
        return
    
    if not context.args:
        await update.message.reply_text("Usage: /makeadmin <user_id>")
        return
    
    target_user_id = context.args[0].strip()
    
    if set_admin_status(target_user_id, True):
        await update.message.reply_text(f"✅ User {target_user_id} is now an admin.")
    else:
        await update.message.reply_text(f"❌ Failed to set admin status for {target_user_id}.")

async def removeadmin_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Owner command to revoke admin access. Usage: /removeadmin <user_id>"""
    if OWNER_CHAT_ID and str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ Owner only command.")
        return
    
    if not context.args:
        await update.message.reply_text("Usage: /removeadmin <user_id>")
        return
    
    target_user_id = context.args[0].strip()
    
    if set_admin_status(target_user_id, False):
        await update.message.reply_text(f"✅ Removed admin status from {target_user_id}.")
    else:
        await update.message.reply_text(f"❌ Failed to remove admin status from {target_user_id}.")

def require_subscription(handler):
    """Decorator to require active subscription for handlers."""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = str(update.effective_user.id)
        has_access, reason = check_access(user_id)
        
        if not has_access:
            if reason == "no_account":
                await update.message.reply_text(
                    "👋 Welcome! Send /start to begin your free trial."
                )
            elif reason == "expired":
                status = get_user_subscription_status(user_id)
                if status and status["trial_active"]:
                    msg = "🎁 Your free trial has ended. Subscribe to continue:\n/subscribe"
                else:
                    msg = "⭐ Your subscription has expired. Renew to continue:\n/subscribe"
                await update.message.reply_text(msg)
            return
        
        # Has access - call the actual handler
        return await handler(update, context)
    
    return wrapper

# ============================================================
# DAILY SUBSCRIPTION CHECK JOB
# ============================================================

async def check_expired_subscriptions(context: ContextTypes.DEFAULT_TYPE):
    """Daily job to notify users about expiring subscriptions."""
    bot = context.bot
    
    query = """
        SELECT user_id, subscription_expiry 
        FROM users 
        WHERE is_subscribed = TRUE 
        AND is_admin = FALSE
        AND subscription_expiry IS NOT NULL
        AND subscription_expiry > CURRENT_TIMESTAMP
        AND subscription_expiry < CURRENT_TIMESTAMP + INTERVAL '3 days';
    """
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()
        cur.close()
        
        log(f"📋 Checking {len(results)} users with expiring subscriptions")
        
        for user_id, expiry in results:
            days_left = (expiry - datetime.now()).days
            
            try:
                await bot.send_message(
                    chat_id=int(user_id),
                    text=(
                        f"⏰ <b>Subscription Reminder</b>\n\n"
                        f"Your subscription expires in {days_left} day(s).\n"
                        f"Renew now to keep your access:\n/subscribe"
                    ),
                    parse_mode="HTML"
                )
                log(f"✅ Sent reminder to {user_id}")
            except Exception as e:
                log(f"❌ Failed to send reminder to {user_id}: {e}")
        
    except Exception as e:
        log(f"❌ Error checking expired subscriptions: {e}")
    finally:
        if conn:
            return_db(conn)

# ============================================================
# ADMIN MESSAGING HANDLERS
# ============================================================

async def messageadmin_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Allow users to send a message to the admin/owner."""
    if not OWNER_CHAT_ID:
        await update.message.reply_text("⚠️ Admin contact is not configured.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /messageadmin <your message>\n\n"
            "Example: /messageadmin I have a question about store tracking"
        )
        return
    
    user = update.effective_user
    user_id = str(user.id)
    username = user.username or "No username"
    full_name = user.full_name or "Unknown"
    
    # Get user's subscription status for context
    status = get_user_subscription_status(user_id)
    if status:
        if status["is_admin"]:
            user_type = "👑 Admin"
        elif status["trial_active"]:
            user_type = "🎁 Trial User"
        elif status["is_active"]:
            user_type = "✅ Subscribed"
        else:
            user_type = "⭐ Expired"
    else:
        user_type = "❓ Unknown"
    
    # Join all arguments to form the message
    user_message = " ".join(context.args)
    
    # Format message to admin
    admin_message = (
        f"📬 <b>Message from User</b>\n\n"
        f"<b>From:</b> {full_name} (@{username})\n"
        f"<b>User ID:</b> <code>{user_id}</code>\n"
        f"<b>Status:</b> {user_type}\n\n"
        f"<b>Message:</b>\n{user_message}\n\n"
        f"<i>Reply with: /reply {user_id} your message</i>"
    )
    
    try:
        # Send to owner
        await context.bot.send_message(
            chat_id=OWNER_CHAT_ID,
            text=admin_message,
            parse_mode="HTML"
        )
        
        # Confirm to user
        await update.message.reply_text(
            "✅ Your message has been sent to the admin. "
            "They will respond as soon as possible."
        )
        
        log(f"📬 Message from user {user_id} sent to admin")
        
    except Exception as e:
        log(f"❌ Error sending message to admin: {e}")
        await update.message.reply_text(
            "⚠️ Failed to send message. Please try again later."
        )


async def reply_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Owner command to reply to a specific user."""
    # Only owner can use this
    if not OWNER_CHAT_ID or str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ Owner only command.")
        return
    
    if len(context.args) < 2:
        await update.message.reply_text(
            "Usage: /reply <user_id> <your message>\n\n"
            "Example: /reply 123456789 Thanks for your message!"
        )
        return
    
    target_user_id = context.args[0]
    reply_message = " ".join(context.args[1:])
    
    # Format message to user
    user_message = (
        f"💬 <b>Message from Admin</b>\n\n"
        f"{reply_message}"
    )
    
    try:
        # Send to user
        await context.bot.send_message(
            chat_id=int(target_user_id),
            text=user_message,
            parse_mode="HTML"
        )
        
        # Confirm to owner
        await update.message.reply_text(
            f"✅ Reply sent to user {target_user_id}"
        )
        
        log(f"💬 Admin reply sent to user {target_user_id}")
        
    except Exception as e:
        log(f"❌ Error sending reply to user {target_user_id}: {e}")
        await update.message.reply_text(
            f"⚠️ Failed to send reply to user {target_user_id}. "
            f"They may have blocked the bot or the ID is invalid."
        )


async def sendall_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Owner command to send a message to all active users."""
    # Only owner can use this
    if not OWNER_CHAT_ID or str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ Owner only command.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /sendall <your message>\n\n"
            "This will send your message to ALL active users (subscribed or trial).\n"
            "Use with caution!"
        )
        return
    
    broadcast_message = " ".join(context.args)
    
    # Get all active users (subscribed, trial, or admin)
    query = """
        SELECT user_id, full_name, is_admin, is_subscribed, subscription_expiry
        FROM users
        WHERE (
            is_admin = TRUE 
            OR (is_subscribed = TRUE AND subscription_expiry > CURRENT_TIMESTAMP)
        );
    """
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        users = cur.fetchall()
        cur.close()
    except Exception as e:
        log(f"❌ Error getting users for broadcast: {e}")
        await update.message.reply_text("⚠️ Database error. Check logs.")
        return
    finally:
        if conn:
            return_db(conn)
    
    if not users:
        await update.message.reply_text("⚠️ No active users found.")
        return
    
    # Confirm before sending
    await update.message.reply_text(
        f"📢 Preparing to send message to {len(users)} active users...\n\n"
        f"<b>Preview:</b>\n{broadcast_message}\n\n"
        f"Sending in 3 seconds...",
        parse_mode="HTML"
    )
    
    await asyncio.sleep(3)
    
    # Format message for users
    user_message = (
        f"📢 <b>Announcement from FWGS Bot</b>\n\n"
        f"{broadcast_message}"
    )
    
    # Send to all users
    success_count = 0
    fail_count = 0
    failed_users = []
    
    for user_id, full_name, is_admin, is_subscribed, expiry in users:
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=user_message,
                parse_mode="HTML"
            )
            success_count += 1
            log(f"✅ Broadcast sent to {user_id} ({full_name})")
            
            # Small delay to avoid rate limits
            await asyncio.sleep(0.05)
            
        except Exception as e:
            fail_count += 1
            failed_users.append(f"{user_id} ({full_name})")
            log(f"❌ Failed to send to {user_id}: {e}")
    
    # Report results
    result_message = (
        f"📊 <b>Broadcast Complete</b>\n\n"
        f"✅ Sent: {success_count}\n"
        f"❌ Failed: {fail_count}"
    )
    
    if failed_users and len(failed_users) <= 10:
        result_message += f"\n\n<b>Failed users:</b>\n" + "\n".join(failed_users)
    
    await update.message.reply_text(result_message, parse_mode="HTML")
    log(f"📢 Broadcast complete: {success_count} sent, {fail_count} failed")


async def sendallwatchlist_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Owner command to send a message to all users who have items in their watchlist."""
    # Only owner can use this
    if not OWNER_CHAT_ID or str(update.effective_user.id) != str(OWNER_CHAT_ID):
        await update.message.reply_text("⛔ Owner only command.")
        return
    
    if not context.args:
        await update.message.reply_text(
            "Usage: /sendallwatchlist <your message>\n\n"
            "This will send your message to ALL users with watchlist items.\n"
            "Good for announcements about monitoring features."
        )
        return
    
    broadcast_message = " ".join(context.args)
    
    # Get all users who have watchlist items
    query = """
        SELECT DISTINCT w.user_id, u.full_name
        FROM watchlist w
        LEFT JOIN users u ON w.user_id = u.user_id
        ORDER BY w.user_id;
    """
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(query)
        users = cur.fetchall()
        cur.close()
    except Exception as e:
        log(f"❌ Error getting watchlist users: {e}")
        await update.message.reply_text("⚠️ Database error. Check logs.")
        return
    finally:
        if conn:
            return_db(conn)
    
    if not users:
        await update.message.reply_text("⚠️ No users with watchlists found.")
        return
    
    # Confirm before sending
    await update.message.reply_text(
        f"📢 Preparing to send message to {len(users)} watchlist users...\n\n"
        f"<b>Preview:</b>\n{broadcast_message}\n\n"
        f"Sending in 3 seconds...",
        parse_mode="HTML"
    )
    
    await asyncio.sleep(3)
    
    # Format message for users
    user_message = (
        f"📢 <b>Announcement from FWGS Bot</b>\n\n"
        f"{broadcast_message}"
    )
    
    # Send to all users
    success_count = 0
    fail_count = 0
    
    for user_id, full_name in users:
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=user_message,
                parse_mode="HTML"
            )
            success_count += 1
            log(f"✅ Broadcast sent to {user_id} ({full_name})")
            await asyncio.sleep(0.05)
            
        except Exception as e:
            fail_count += 1
            log(f"❌ Failed to send to {user_id}: {e}")
    
    # Report results
    await update.message.reply_text(
        f"📊 <b>Broadcast Complete</b>\n\n"
        f"✅ Sent: {success_count}\n"
        f"❌ Failed: {fail_count}",
        parse_mode="HTML"
    )
    log(f"📢 Watchlist broadcast complete: {success_count} sent, {fail_count} failed")


# ============================================================
# SCHEDULER & RUNNER
# ============================================================

main_loop = None

def schedule_coroutine(coro_func, *args, **kwargs):
    """Schedule an async coroutine to run in the main asyncio loop."""
    if main_loop is None:
        raise RuntimeError("Main loop not set; cannot schedule coroutine.")
    future = asyncio.run_coroutine_threadsafe(
        coro_func(*args, **kwargs), 
        main_loop
    )
    def _cb(fut):
        try:
            fut.result()
        except Exception as e:
            log(f"❌ Exception in scheduled job {getattr(coro_func, '__name__', str(coro_func))}: {e}")
    future.add_done_callback(_cb)
    return future

async def runner():
    """Main runner function."""
    global main_loop
    main_loop = asyncio.get_running_loop()
    
    # Initialize database
    init_connection_pool()
    init_db()
    
    # Build application
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    log("✅ App built successfully")

    app.add_error_handler(error_handler)
    
    # Initialize app
    await app.initialize()
    
    # Register handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add", add_product))
    app.add_handler(CommandHandler("remove", remove_product))
    app.add_handler(CommandHandler("watchlist", show_watchlist))
    app.add_handler(CommandHandler("status", status))
    app.add_handler(CommandHandler("kill", kill_bot))
    app.add_handler(CommandHandler("global", show_global_list))
    app.add_handler(CommandHandler("lastreport", send_global_report))
    app.add_handler(CommandHandler("addstore", addstore_handler))
    app.add_handler(CommandHandler("removestore", removestore_handler))
    app.add_handler(CommandHandler("mystores", mystores_handler))
    app.add_handler(CommandHandler("statestock", statestock_handler))
    app.add_handler(CommandHandler("testreport", test_report))
    app.add_handler(CommandHandler("subscribe", subscribe_handler))
    app.add_handler(CommandHandler("status", status_handler))
    app.add_handler(PreCheckoutQueryHandler(precheckout_handler))
    app.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment_handler))
    app.add_handler(CommandHandler("makeadmin", makeadmin_handler))
    app.add_handler(CommandHandler("removeadmin", removeadmin_handler))
    app.add_handler(CommandHandler("messageadmin", messageadmin_handler))
    app.add_handler(CommandHandler("reply", reply_handler))
    app.add_handler(CommandHandler("sendall", sendall_handler))
    app.add_handler(CommandHandler("sendallwatchlist", sendallwatchlist_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    log("✅ Handlers registered")
    
    # Start app
    await app.start()
    
    # JobQueue tasks
    app.job_queue.run_repeating(active_monitor, interval=30, first=10)
    app.job_queue.run_repeating(category_monitor, interval=300, first=15)
    app.job_queue.run_repeating(inventory_refresh_job, interval=1800, first=10)
    
        # APScheduler
    scheduler = AsyncIOScheduler(job_defaults={"misfire_grace_time": 60})
    scheduler.add_job(
        lambda: schedule_coroutine(refresh_global_list, app),
        trigger="cron",
        hour=9,
        minute=0,
    )
    
    import warnings
    from telegram.warnings import PTBUserWarning
    warnings.filterwarnings("ignore", category=PTBUserWarning)
    
    # Daily subscription check (9 AM)
    app.job_queue.run_daily(
        check_expired_subscriptions,
        time=time(hour=8, minute=0),
        days=(0, 1, 2, 3, 4, 5, 6)
    )

    scheduler.start()
    log("⏰ Scheduler started and attached to main event loop")
    
    # Start polling
    log("🤖 Bot is running with scheduler active...")
    await app.updater.start_polling()
    
    try:
        # Keep running until interrupted
        await asyncio.Event().wait()
    except (KeyboardInterrupt, SystemExit):
        log("🛑 Bot stopped manually")
    finally:
        # Cleanup
        await app.updater.stop()
        await app.stop()
        await app.shutdown()
        close_pool()
        log("✅ Bot shutdown complete")

# ============================================================
# ENTRY POINT
# ============================================================

def main():
    """Entry point that handles event loop properly."""
    try:
        asyncio.run(runner())
    except KeyboardInterrupt:
        log("🛑 Stopped by user")
    except Exception as e:
        log(f"❌ Fatal error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
